"""Import de questions de quiz depuis un fichier (CSV / Excel .xlsx / JSON).

Format tabulaire (CSV/Excel) — 1 question par ligne, en-têtes (insensibles à la casse) :
    question | option1 | option2 | … | option6 | correct | points | type
  - question : intitulé (obligatoire)
  - option1..option6 : réponses possibles (≥ 2 non vides)
  - correct : n° des bonnes réponses en base 1 ("2" ou "1,3") ou lettres ("B", "A,C")
  - points  : entier (défaut 1)
  - type    : "radio"/"unique" (une réponse) ou "checkbox"/"multiple" (plusieurs).
              Déduit automatiquement si absent (radio si 1 bonne réponse, sinon checkbox).

Format JSON — liste d'objets {title, description?, points?, input_type?/type?,
options:[{title, is_answer|correct}]} (ou {"questions": [...]}).

`parse_quiz_file` renvoie (questions, erreurs) : chaque question au format « à plat »
attendu par QuizQuestionViewSet ({title, description, points, input_type, options}).
"""
import csv
import io
import json

MAX_OPTIONS = 6
TEMPLATE_HEADER = ["question", "option1", "option2", "option3", "option4", "correct", "points", "type"]
TEMPLATE_SAMPLE = [
    ["Quelle est la capitale du Cameroun ?", "Douala", "Yaoundé", "Kribi", "Bafoussam", "2", "1", "radio"],
    ["Lesquels sont des langages de programmation ?", "Python", "HTML", "Java", "CSS", "1,3", "2", "checkbox"],
]


def _norm(v):
    return ("" if v is None else str(v)).strip()


def _parse_correct(raw, n_options):
    """« 2 », « 1,3 », « B », « A;C » → ensemble d'index en base 1."""
    out = set()
    for tok in _norm(raw).replace(";", ",").replace(" ", ",").split(","):
        tok = tok.strip()
        if not tok:
            continue
        if tok.isdigit():
            idx = int(tok)
        elif len(tok) == 1 and tok.isalpha():
            idx = ord(tok.upper()) - ord("A") + 1
        else:
            continue
        if 1 <= idx <= n_options:
            out.add(idx)
    return out


def _resolve_type(raw, n_correct):
    t = _norm(raw).lower()
    if t in ("checkbox", "multiple", "multi", "1", "cases"):
        return 1
    if t in ("radio", "unique", "single", "2", "qcu"):
        return 2
    return 1 if n_correct > 1 else 2  # déduction


def _row_to_question(row):
    title = _norm(row.get("question"))
    if not title:
        return None, None  # ligne vide → ignorée sans erreur
    options = [o for i in range(1, MAX_OPTIONS + 1) if (o := _norm(row.get(f"option{i}")))]
    if len(options) < 2:
        return None, f"« {title[:40]} » : au moins 2 options requises."
    correct = _parse_correct(row.get("correct") or row.get("correcte") or row.get("reponse"), len(options))
    if not correct:
        return None, f"« {title[:40]} » : aucune bonne réponse indiquée (colonne « correct »)."
    try:
        points = max(1, int(float(_norm(row.get("points")) or 1)))
    except (TypeError, ValueError):
        points = 1
    return {
        "title": title,
        "description": _norm(row.get("description")),
        "points": points,
        "input_type": _resolve_type(row.get("type"), len(correct)),
        "options": [{"title": o, "is_answer": (i + 1) in correct} for i, o in enumerate(options)],
    }, None


def _rows_from_csv(data):
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:2000]
    delim = ";" if sample.count(";") > sample.count(",") else ","
    for r in csv.DictReader(io.StringIO(text), delimiter=delim):
        yield {(k or "").strip().lower(): v for k, v in r.items()}


def _rows_from_xlsx(data):
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    header = [(_norm(c).lower()) for c in next(it, [])]
    for row in it:
        yield {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}


def _questions_from_json(data):
    obj = json.loads(data.decode("utf-8-sig", errors="replace"))
    items = obj.get("questions") if isinstance(obj, dict) else obj
    out, errors = [], []
    for it in (items or []):
        if not isinstance(it, dict):
            continue
        title = _norm(it.get("title") or it.get("question"))
        if not title:
            continue
        opts = []
        for o in (it.get("options") or []):
            if isinstance(o, dict):
                t = _norm(o.get("title") or o.get("text"))
                if t:
                    opts.append({"title": t, "is_answer": bool(o.get("is_answer") or o.get("correct"))})
            elif _norm(o):
                opts.append({"title": _norm(o), "is_answer": False})
        n_correct = sum(1 for o in opts if o["is_answer"])
        if len(opts) < 2 or n_correct == 0:
            errors.append(f"« {title[:40]} » : 2 options min. et au moins une bonne réponse.")
            continue
        try:
            points = max(1, int(it.get("points", 1)))
        except (TypeError, ValueError):
            points = 1
        out.append({
            "title": title,
            "description": _norm(it.get("description")),
            "points": points,
            "input_type": _resolve_type(it.get("input_type") or it.get("type"), n_correct),
            "options": opts,
        })
    return out, errors


def parse_quiz_file(filename, data):
    """(questions, erreurs) depuis le contenu d'un fichier CSV / .xlsx / .json."""
    name = (filename or "").lower()
    if name.endswith(".json"):
        return _questions_from_json(data)
    rows = _rows_from_xlsx(data) if name.endswith(".xlsx") else _rows_from_csv(data)
    out, errors = [], []
    for r in rows:
        q, err = _row_to_question(r)
        if err:
            errors.append(err)
        elif q:
            out.append(q)
    return out, errors


def build_template(fmt):
    """(bytes, content_type, filename) d'un gabarit d'import prérempli."""
    if fmt == "xlsx":
        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "Quiz"
        ws.append(TEMPLATE_HEADER)
        for r in TEMPLATE_SAMPLE:
            ws.append(r)
        bio = io.BytesIO()
        wb.save(bio)
        return (
            bio.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "modele_quiz.xlsx",
        )
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(TEMPLATE_HEADER)
    w.writerows(TEMPLATE_SAMPLE)
    return buf.getvalue().encode("utf-8-sig"), "text/csv; charset=utf-8", "modele_quiz.csv"
